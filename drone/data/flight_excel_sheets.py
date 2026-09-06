from __future__ import annotations

import math
import time
from typing import Callable, Optional

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from drone.data.flight_excel_style import (
    ALERT_FILL,
    ALERT_FONT,
    BASE_FONT,
    CELL_BORDER,
    CENTER,
    FMT_DEG,
    FMT_INT,
    FMT_M,
    FMT_PCT,
    FMT_S,
    FMT_TIME,
    HEADER_BORDER,
    HEADER_FILL,
    HEADER_FONT,
    LABEL_FONT,
    LEFT,
    OK_FILL,
    OK_FONT,
    RIGHT,
    SUBTITLE_FONT,
    TITLE_FONT,
    WARN_FILL,
    WARN_FONT,
    WRAP_LEFT,
    ZEBRA_FILL,
    autofit_columns,
    bool_it,
    clock,
    fmt_num_it,
    kv,
    section_title,
    table_header,
    translate_reason,
    translate_source,
)
from drone.data.flight_report_stats import (
    elapsed_seconds,
    filter_divergence,
    join_tag_ids,
    percent_within,
    rms,
    total_duration_sec,
    values_of,
    waypoint_groups,
)
from drone.geometry import circular_mean_deg


def _std(values: list[float], mean_value: float) -> Optional[float]:
    if not values:
        return None
    return math.sqrt(sum((v - mean_value) ** 2 for v in values) / len(values))


def _circular_std_deg(values: list[float]) -> Optional[float]:
    if not values:
        return None
    rad = [math.radians(v) for v in values]
    cos_sum = sum(math.cos(r) for r in rad) / len(rad)
    sin_sum = sum(math.sin(r) for r in rad) / len(rad)
    resultant = math.hypot(cos_sum, sin_sum)
    if resultant <= 0.0:
        return None
    return math.degrees(math.sqrt(max(0.0, -2.0 * math.log(resultant))))


def _column_stats(entries, columns: list[str]) -> list[tuple[str, float, float, float, Optional[float]]]:
    out: list[tuple[str, float, float, float, Optional[float]]] = []
    for col in columns:
        values = [float(e[col]) for e in entries if e.get(col) is not None]
        if not values:
            continue
        if "yaw" in col and "error" not in col:
            mean_value = circular_mean_deg(values)
            std_value = _circular_std_deg(values)
        else:
            mean_value = sum(values) / len(values)
            std_value = _std(values, mean_value)
        out.append((col, min(values), max(values), mean_value, std_value))
    return out


def _filter_divergence_flags(entries, margin_m: float) -> list[str]:
    return [
        f"Asse {d['axis']}: filtrata [{d['filtered_min']:.3f}, {d['filtered_max']:.3f}] m "
        f"oltre grezza [{d['raw_min']:.3f}, {d['raw_max']:.3f}] m di {d['excess']:.3f} m"
        for d in filter_divergence(entries, margin_m)
    ]


def _rel_time(entries):
    tempi = {id(e): ti for e, ti in zip(entries, elapsed_seconds(entries))}
    return lambda e: tempi.get(id(e), 0.0)


def _autopilot_columns(entries) -> list[tuple[str, Callable, Optional[str]]]:
    rel = _rel_time(entries)

    def wp_number(e):
        idx = e.get("target_index")
        return None if idx is None else int(idx) + 1

    return [
        ("Orario", lambda e: clock(e["timestamp"]), FMT_TIME),
        ("Tempo dall'avvio [s]", rel, FMT_S),
        ("Drone X [m]", lambda e: e["x"], FMT_M),
        ("Drone Y [m]", lambda e: e["y"], FMT_M),
        ("Drone Z [m]", lambda e: e["z"], FMT_M),
        ("Drone yaw [°]", lambda e: e["yaw_deg"], FMT_DEG),
        ("Waypoint X [m]", lambda e: e.get("target_x"), FMT_M),
        ("Waypoint Y [m]", lambda e: e.get("target_y"), FMT_M),
        ("Waypoint Z [m]", lambda e: e.get("target_z"), FMT_M),
        ("Waypoint yaw [°]", lambda e: e.get("target_yaw_deg"), FMT_DEG),
        ("Comando avanti/indietro", lambda e: e["fb"], FMT_INT),
        ("Comando destra/sinistra", lambda e: e["lr"], FMT_INT),
        ("Comando salita/discesa", lambda e: e["ud"], FMT_INT),
        ("Comando rotazione", lambda e: e["yaw_cmd"], FMT_INT),
        ("Distanza orizzontale [m]", lambda e: e.get("distance_xy"), FMT_M),
        ("Distanza nello spazio [m]", lambda e: e.get("distance_3d"), FMT_M),
        ("Errore orientamento [°]", lambda e: e.get("yaw_error_deg"), FMT_DEG),
        ("Waypoint attivo (n.)", wp_number, FMT_INT),
        ("Waypoint raggiunto", lambda e: bool_it(e.get("reached")), None),
        ("Missione completata", lambda e: bool_it(e.get("finished")), None),
        ("Anomalia", lambda e: bool_it(e.get("fault")), None),
        ("Stato del controllo", lambda e: translate_reason(e.get("reason", "")), None),
        ("Tag AprilTag visti", lambda e: None, None),
        ("Origine della posa", lambda e: translate_source(e.get("pose_source", "")), None),
    ]


def _comparison_columns(entries) -> list[tuple[str, Callable, Optional[str]]]:
    rel = _rel_time(entries)
    return [
        ("Orario", lambda e: clock(e["timestamp"]), FMT_TIME),
        ("Tempo dall'avvio [s]", rel, FMT_S),
        ("Grezza X [m]", lambda e: e["raw_x"], FMT_M),
        ("Grezza Y [m]", lambda e: e["raw_y"], FMT_M),
        ("Grezza Z [m]", lambda e: e["raw_z"], FMT_M),
        ("Grezza yaw [°]", lambda e: e["raw_yaw_deg"], FMT_DEG),
        ("Kalman X [m]", lambda e: e["filtered_x"], FMT_M),
        ("Kalman Y [m]", lambda e: e["filtered_y"], FMT_M),
        ("Kalman Z [m]", lambda e: e["filtered_z"], FMT_M),
        ("Kalman yaw [°]", lambda e: e["filtered_yaw_deg"], FMT_DEG),
        ("Correzione X [m]", lambda e: e["error_x"], FMT_M),
        ("Correzione Y [m]", lambda e: e["error_y"], FMT_M),
        ("Correzione Z [m]", lambda e: e["error_z"], FMT_M),
        ("Correzione totale [m]", lambda e: e["error_norm"], FMT_M),
        ("Tag AprilTag visti", lambda e: None, None),
        ("Origine posa grezza", lambda e: translate_source(e.get("raw_source")), None),
        ("Origine posa Kalman", lambda e: translate_source(e.get("filtered_source")), None),
        ("Misura scartata", lambda e: bool_it(e.get("outlier_rejected")), None),
    ]


def _write_data_sheet(
    ws: Worksheet,
    entries,
    columns: list[tuple[str, Callable, Optional[str]]],
    *,
    separator: str,
    tag_column_header: str,
    highlights: dict[str, tuple[str, PatternFill, Font]],
) -> None:
    headers = [c[0] for c in columns]
    tag_col_index = headers.index(tag_column_header) + 1

    for col_index, (header, _getter, _fmt) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_index, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = HEADER_BORDER

    for row_index, entry in enumerate(entries, start=2):
        zebra = ZEBRA_FILL if (row_index % 2 == 0) else None
        for col_index, (header, getter, fmt) in enumerate(columns, start=1):
            if col_index == tag_col_index:
                value = join_tag_ids(entry.get("tag_ids"), separator)
            else:
                value = getter(entry)
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.font = BASE_FONT
            if fmt is not None:
                cell.number_format = fmt
                cell.alignment = RIGHT
            else:
                cell.alignment = LEFT
            if zebra is not None:
                cell.fill = zebra

        for header, (trigger_word, fill, font) in highlights.items():
            hi_index = headers.index(header) + 1
            hi_cell = ws.cell(row=row_index, column=hi_index)
            if str(hi_cell.value) == trigger_word:
                hi_cell.fill = fill
                hi_cell.font = font
                hi_cell.alignment = CENTER

    autofit_columns(ws)

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(columns))
    last_row = len(entries) + 1
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"


_STAT_LABELS = {
    "x": "Drone X [m]", "y": "Drone Y [m]", "z": "Drone Z [m]", "yaw_deg": "Drone yaw [°]",
    "distance_xy": "Distanza orizzontale [m]", "distance_3d": "Distanza nello spazio [m]",
    "yaw_error_deg": "Errore orientamento [°]",
    "raw_x": "Grezza X [m]", "raw_y": "Grezza Y [m]", "raw_z": "Grezza Z [m]",
    "raw_yaw_deg": "Grezza yaw [°]",
    "filtered_x": "Kalman X [m]", "filtered_y": "Kalman Y [m]",
    "filtered_z": "Kalman Z [m]", "filtered_yaw_deg": "Kalman yaw [°]",
    "error_norm": "Correzione totale [m]",
}


def _write_stats_table(ws: Worksheet, row: int, stats: list[tuple[str, float, float, float, Optional[float]]]) -> int:
    row = table_header(ws, row, ["Grandezza", "Minimo", "Massimo", "Media", "Dev. std"])
    for name, mn, mx, mean, std in stats:
        label = _STAT_LABELS.get(name, name)
        fmt = FMT_DEG if "yaw" in name else FMT_M
        ws.cell(row=row, column=1, value=label).font = BASE_FONT
        for col_index, val in ((2, mn), (3, mx), (4, mean), (5, std)):
            if val is None:
                c = ws.cell(row=row, column=col_index, value="—")
                c.alignment = CENTER
            else:
                c = ws.cell(row=row, column=col_index, value=val)
                c.number_format = fmt
                c.alignment = RIGHT
            c.font = BASE_FONT
            c.border = CELL_BORDER
        ws.cell(row=row, column=1).border = CELL_BORDER
        ws.cell(row=row, column=1).alignment = LEFT
        row += 1
    return row


def write_summary_sheet(ws: Worksheet, logger) -> None:
    ws.sheet_view.showGridLines = False

    auto = list(logger.autopilot_entries)
    comp = list(logger.comparison_entries)

    ws.merge_cells("A1:E1")
    title = ws.cell(row=1, column=1, value="Riepilogo sessione di volo")
    title.font = TITLE_FONT
    ws.merge_cells("A2:E2")
    subtitle = ws.cell(
        row=2, column=1,
        value=f"Generato il {time.strftime('%d/%m/%Y alle %H:%M:%S')}",
    )
    subtitle.font = SUBTITLE_FONT
    row = 4

    row = section_title(ws, row, "Panoramica", span=5)
    if auto:
        t0, t1 = float(auto[0]["timestamp"]), float(auto[-1]["timestamp"])
        duration = total_duration_sec(auto)
        row = kv(ws, row, "Campioni registrati (volo autonomo)", len(auto), FMT_INT)
        row = kv(ws, row, "Durata volo autonomo [s]", duration, "0.0")
        if duration > 0:
            row = kv(ws, row, "Frequenza media di campionamento [Hz]", (len(auto) - 1) / duration, "0.0")
        row = kv(
            ws, row, "Orario di volo",
            f"da {time.strftime('%H:%M:%S', time.localtime(t0))} "
            f"a {time.strftime('%H:%M:%S', time.localtime(t1))}",
        )
    if comp:
        row = kv(ws, row, "Campioni confronto grezza/Kalman", len(comp), FMT_INT)
        if not auto:
            row = kv(ws, row, "Durata registrazione [s]", total_duration_sec(comp), "0.0")
    row += 1

    if auto:
        groups = waypoint_groups(auto)
        indices = sorted({g["index"] for g in groups if g["index"] is not None})
        reached_indices = {g["index"] for g in groups if g["reached"] and g["index"] is not None}
        completed = any(e.get("finished") for e in auto)
        fault_entry = next((e for e in reversed(auto) if e.get("fault")), None)

        row = section_title(ws, row, "Esito della missione", span=5)
        row = kv(ws, row, "Waypoint raggiunti", f"{len(reached_indices & set(indices))} su {len(indices)}")
        row = kv(ws, row, "Missione completata", "Sì" if completed else "No")
        if fault_entry is not None:
            row = kv(ws, row, "Anomalia rilevata", translate_reason(fault_entry.get("reason", "")))
        row += 1

    if auto:
        row = section_title(ws, row, "Qualità del volo autonomo", span=5)
        rms_xy = rms(values_of(auto, "distance_xy"))
        rms_3d = rms(values_of(auto, "distance_3d"))
        rms_yaw = rms(values_of(auto, "yaw_error_deg"))
        if rms_xy is not None:
            row = kv(ws, row, "Errore medio orizzontale (RMS) [m]", rms_xy, FMT_M)
        if rms_3d is not None:
            row = kv(ws, row, "Errore medio nello spazio (RMS) [m]", rms_3d, FMT_M)
        if rms_yaw is not None:
            row = kv(ws, row, "Errore medio di orientamento (RMS) [°]", rms_yaw, FMT_DEG)

        xy_tol = logger.autopilot_xy_tolerance_m
        z_tol = logger.autopilot_z_tolerance_m
        yaw_tol = logger.autopilot_yaw_tolerance_deg
        if xy_tol is not None:
            pct = percent_within(
                auto,
                lambda e: float(e["distance_xy"]) <= xy_tol,
                lambda e: e.get("distance_xy") is not None,
            )
            if pct is not None:
                row = kv(ws, row, f"Tempo entro tolleranza orizzontale (≤ {fmt_num_it(xy_tol)} m)", pct, FMT_PCT)
        if z_tol is not None:
            pct = percent_within(
                auto,
                lambda e: abs(float(e["z"]) - float(e["target_z"])) <= z_tol,
                lambda e: e.get("z") is not None and e.get("target_z") is not None,
            )
            if pct is not None:
                row = kv(ws, row, f"Tempo entro tolleranza di quota (≤ {fmt_num_it(z_tol)} m)", pct, FMT_PCT)
        if yaw_tol is not None:
            pct = percent_within(
                auto,
                lambda e: abs(float(e["yaw_error_deg"])) <= yaw_tol,
                lambda e: e.get("yaw_error_deg") is not None,
            )
            if pct is not None:
                row = kv(ws, row, f"Tempo entro tolleranza di orientamento (≤ {fmt_num_it(yaw_tol, 1)} °)", pct, FMT_PCT)
        row += 1

        row = section_title(ws, row, "Tempo per waypoint", span=5)
        row = table_header(ws, row, ["Waypoint", "Durata [s]", "Distanza minima [m]", "Raggiunto"])
        for g in waypoint_groups(auto):
            label = "—" if g["index"] is None else str(int(g["index"]) + 1)
            duration = g["end"] - g["start"]
            wp_cell = ws.cell(row=row, column=1, value=label)
            wp_cell.font = BASE_FONT
            wp_cell.alignment = CENTER
            wp_cell.border = CELL_BORDER
            dur_cell = ws.cell(row=row, column=2, value=duration)
            dur_cell.font = BASE_FONT
            dur_cell.number_format = "0.0"
            dur_cell.alignment = RIGHT
            dur_cell.border = CELL_BORDER
            if g["min_d3"] is None:
                md_cell = ws.cell(row=row, column=3, value="—")
                md_cell.alignment = CENTER
            else:
                md_cell = ws.cell(row=row, column=3, value=g["min_d3"])
                md_cell.number_format = FMT_M
                md_cell.alignment = RIGHT
            md_cell.font = BASE_FONT
            md_cell.border = CELL_BORDER
            reached = bool(g["reached"])
            re_cell = ws.cell(row=row, column=4, value=bool_it(reached))
            re_cell.alignment = CENTER
            re_cell.border = CELL_BORDER
            if reached:
                re_cell.fill = OK_FILL
                re_cell.font = OK_FONT
            else:
                re_cell.fill = ALERT_FILL
                re_cell.font = ALERT_FONT
            row += 1
        row += 1

        row = section_title(ws, row, "Valori registrati (autopilota)", span=5)
        row = _write_stats_table(
            ws, row,
            _column_stats(auto, ["x", "y", "z", "yaw_deg", "distance_xy", "distance_3d", "yaw_error_deg"]),
        )
        row += 1

    if comp:
        row = section_title(ws, row, "Confronto grezza/filtrata (Kalman)", span=5)
        row = _write_stats_table(
            ws, row,
            _column_stats(comp, [
                "raw_x", "raw_y", "raw_z", "raw_yaw_deg",
                "filtered_x", "filtered_y", "filtered_z", "filtered_yaw_deg",
                "error_norm",
            ]),
        )
        row += 1

        flags = _filter_divergence_flags(comp, logger.FILTER_DIVERGENCE_MARGIN_M)
        if flags:
            row = section_title(ws, row, "Avviso: divergenza del filtro Kalman", span=5)
            note = ws.cell(
                row=row, column=1,
                value=(
                    f"La posa filtrata esce dal range della grezza oltre "
                    f"{fmt_num_it(logger.FILTER_DIVERGENCE_MARGIN_M, 2)} m. Possibile correzione "
                    "divergente o outlier non scartato dal gate: verifica "
                    "process_noise, measurement_noise e la soglia del gate."
                ),
            )
            note.font = BASE_FONT
            note.alignment = WRAP_LEFT
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            ws.row_dimensions[row].height = 44
            row += 1
            for item in flags:
                c = ws.cell(row=row, column=1, value=f"• {item}")
                c.font = BASE_FONT
                c.alignment = LEFT
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
                row += 1

    autofit_columns(ws)


_LEGEND_AUTOPILOT = [
    ("Orario", "Ora dell'orologio in cui è stato registrato il campione."),
    ("Tempo dall'avvio [s]", "Secondi trascorsi dal primo campione del volo autonomo."),
    ("Drone X / Y / Z [m]", "Posizione del drone stimata dalla visione, quella usata dall'autopilota."),
    ("Drone yaw [°]", "Orientamento del drone (yaw = imbardata, rotazione attorno all'asse verticale)."),
    ("Waypoint X / Y / Z [m]", "Posizione del punto di passaggio verso cui il drone si dirige."),
    ("Waypoint yaw [°]", "Orientamento desiderato una volta arrivati al waypoint."),
    ("Comando avanti/indietro", "Spinta longitudinale ai motori (positivo = avanti, negativo = indietro)."),
    ("Comando destra/sinistra", "Spinta laterale ai motori (positivo = destra, negativo = sinistra)."),
    ("Comando salita/discesa", "Comando di quota ai motori (positivo = salita, negativo = discesa)."),
    ("Comando rotazione", "Comando di rotazione su sé stesso (positivo = senso orario)."),
    ("Distanza orizzontale [m]", "Distanza dal waypoint sul piano, senza contare la quota."),
    ("Distanza nello spazio [m]", "Distanza dal waypoint considerando anche la quota."),
    ("Errore orientamento [°]", "Quanto l'orientamento attuale è lontano da quello desiderato."),
    ("Waypoint attivo (n.)", "Numero del waypoint verso cui il drone si sta dirigendo (a partire da 1)."),
    ("Waypoint raggiunto", "Sì quando il waypoint è stato considerato raggiunto in quel momento."),
    ("Missione completata", "Sì quando tutta la sequenza di waypoint è stata percorsa."),
    ("Anomalia", "Sì quando il controllo segnala un problema (posizione persa o tempo scaduto sul waypoint)."),
    ("Stato del controllo", "Cosa sta facendo l'autopilota (inseguimento, sosta di supervisione, correzione di quota...)."),
    ("Tag AprilTag visti", "Numeri degli AprilTag usati per calcolare la posizione in quel campione."),
    ("Origine della posa", "Da dove viene la posizione usata (es. filtro Kalman sulla media pesata dei tag)."),
]


_LEGEND_KALMAN = [
    ("Orario", "Ora dell'orologio in cui è stato registrato il campione."),
    ("Tempo dall'avvio [s]", "Secondi trascorsi dal primo campione registrato."),
    ("Grezza X / Y / Z [m]", "Posizione misurata direttamente dagli AprilTag, senza filtro."),
    ("Grezza yaw [°]", "Orientamento misurato direttamente dagli AprilTag."),
    ("Kalman X / Y / Z [m]", "Posizione dopo il filtro di Kalman, che riduce il rumore della misura."),
    ("Kalman yaw [°]", "Orientamento dopo il filtro (se attivo; altrimenti uguale alla grezza)."),
    ("Correzione X / Y / Z [m]", "Quanto il filtro ha spostato la misura su ciascun asse (Kalman − grezza)."),
    ("Correzione totale [m]", "Entità complessiva della correzione introdotta dal filtro."),
    ("Tag AprilTag visti", "Numeri degli AprilTag usati per la stima in quel campione."),
    ("Origine posa grezza", "Da dove viene la posa grezza."),
    ("Origine posa Kalman", "Da dove viene la posa filtrata."),
    ("Misura scartata", "Sì quando il filtro ha rifiutato la misura grezza perché anomala (outlier)."),
]


def write_legend_sheet(ws: Worksheet, *, include_autopilot: bool, include_kalman: bool) -> None:
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 78
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:B1")
    title = ws.cell(row=1, column=1, value="Legenda delle colonne")
    title.font = TITLE_FONT
    row = 3

    def _block(section_name: str, pairs: list[tuple[str, str]], start: int) -> int:
        r = section_title(ws, start, section_name, span=2)
        r = table_header(ws, r, ["Colonna", "Significato"])
        for name, meaning in pairs:
            name_cell = ws.cell(row=r, column=1, value=name)
            name_cell.font = LABEL_FONT
            name_cell.alignment = WRAP_LEFT
            name_cell.border = CELL_BORDER
            mean_cell = ws.cell(row=r, column=2, value=meaning)
            mean_cell.font = BASE_FONT
            mean_cell.alignment = WRAP_LEFT
            mean_cell.border = CELL_BORDER
            r += 1
        return r + 1

    if include_autopilot:
        row = _block("Foglio «Autopilota»", _LEGEND_AUTOPILOT, row)
    if include_kalman:
        row = _block("Foglio «Kalman»", _LEGEND_KALMAN, row)


def write_parameters_sheet(ws: Worksheet, parameters: list[tuple[str, list[tuple[str, str]]]]) -> None:
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:B1")
    title = ws.cell(row=1, column=1, value="Parametri della sessione")
    title.font = TITLE_FONT
    ws.merge_cells("A2:B2")
    subtitle = ws.cell(
        row=2, column=1,
        value="Configurazione con cui è stato eseguito il volo.",
    )
    subtitle.font = SUBTITLE_FONT
    row = 4

    for group_title, items in parameters:
        row = section_title(ws, row, group_title, span=2)
        for label, value in items:
            label_cell = ws.cell(row=row, column=1, value=label)
            label_cell.font = LABEL_FONT
            label_cell.alignment = LEFT
            label_cell.border = CELL_BORDER
            value_cell = ws.cell(row=row, column=2, value=value)
            value_cell.font = BASE_FONT
            value_cell.alignment = LEFT
            value_cell.border = CELL_BORDER
            row += 1
        row += 1

    note = ws.cell(
        row=row, column=1,
        value="I comandi ai motori sono espressi in unità del canale RC del Tello (intervallo −100…100).",
    )
    note.font = SUBTITLE_FONT
    note.alignment = WRAP_LEFT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    autofit_columns(ws)


def write_autopilot_sheet(ws: Worksheet, entries, *, separator: str) -> None:
    _write_data_sheet(
        ws,
        entries,
        _autopilot_columns(entries),
        separator=separator,
        tag_column_header="Tag AprilTag visti",
        highlights={
            "Anomalia": ("Sì", ALERT_FILL, ALERT_FONT),
            "Waypoint raggiunto": ("Sì", OK_FILL, OK_FONT),
        },
    )


def write_kalman_sheet(ws: Worksheet, entries, *, separator: str) -> None:
    _write_data_sheet(
        ws,
        entries,
        _comparison_columns(entries),
        separator=separator,
        tag_column_header="Tag AprilTag visti",
        highlights={
            "Misura scartata": ("Sì", WARN_FILL, WARN_FONT),
        },
    )
