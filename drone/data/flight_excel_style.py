from __future__ import annotations

from datetime import datetime
from typing import Any, Optional

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet


_FONT_NAME = "Arial"

TITLE_FONT = Font(name=_FONT_NAME, size=15, bold=True, color="1F2937")
SUBTITLE_FONT = Font(name=_FONT_NAME, size=10, italic=True, color="4B5563")
SECTION_FONT = Font(name=_FONT_NAME, size=11, bold=True, color="FFFFFF")
HEADER_FONT = Font(name=_FONT_NAME, size=10, bold=True, color="FFFFFF")
LABEL_FONT = Font(name=_FONT_NAME, size=10, bold=True, color="1F2937")
BASE_FONT = Font(name=_FONT_NAME, size=10, color="1F2937")

HEADER_FILL = PatternFill("solid", fgColor="2F3B52")
SECTION_FILL = PatternFill("solid", fgColor="4A5A78")
ZEBRA_FILL = PatternFill("solid", fgColor="F3F5F8")

OK_FILL = PatternFill("solid", fgColor="CDE9CD")
ALERT_FILL = PatternFill("solid", fgColor="F6C9C4")
WARN_FILL = PatternFill("solid", fgColor="FCE3B4")
OK_FONT = Font(name=_FONT_NAME, size=10, bold=True, color="1B5E20")
ALERT_FONT = Font(name=_FONT_NAME, size=10, bold=True, color="8A1C13")
WARN_FONT = Font(name=_FONT_NAME, size=10, bold=True, color="7A4E00")

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
WRAP_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

_THIN = Side(style="thin", color="D0D5DD")
HEADER_BORDER = Border(bottom=Side(style="medium", color="1F2937"))
CELL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

FMT_M = "0.000"
FMT_DEG = "0.0"
FMT_S = "0.000"
FMT_PCT = '0.0"%"'
FMT_TIME = "hh:mm:ss"
FMT_INT = "0"

_COL_MIN_WIDTH = 8
_COL_MAX_WIDTH = 90
_COL_PADDING = 4


def bool_it(value: Any) -> str:
    return "Sì" if bool(value) else "No"


def clock(timestamp: float) -> datetime:
    return datetime.fromtimestamp(float(timestamp))


_REASON_MAP = {
    "idle": "In attesa",
    "tracking": "Inseguimento waypoint",
    "z_priority_tracking": "Correzione di quota",
    "waypoint_reached": "Waypoint raggiunto",
    "waypoint_timeout": "Tempo scaduto sul waypoint",
    "mission_finished": "Missione completata",
    "pose_timeout": "Posizione persa",
    "pose_missing": "Posizione assente",
    "home": "Rientro alla base",
    "supervision_stop": "Sosta di supervisione",
    "supervision_stop_started": "Supervisione avviata",
    "supervision_stop_completed": "Supervisione completata",
    "supervision_stop_tracking": "Supervisione: assestamento",
    "supervision_stop_z_priority_tracking": "Supervisione: quota",
}


def _renamed_reason(raw: str) -> str:
    return raw.replace("supervision_hold", "supervision_stop")


def translate_reason(raw: Any) -> str:
    if raw is None:
        return ""
    s = _renamed_reason(str(raw).strip())
    if s in _REASON_MAP:
        return _REASON_MAP[s]
    if s.startswith("invalid_pose_transient"):
        return "Posizione non valida (temporanea)"
    if s.startswith("invalid_pose"):
        return "Posizione non valida"
    return s


_SOURCE_BASE = {
    "weighted_average": "media pesata dei tag",
    "best_tag": "tag migliore",
    "raw": "grezza",
}


def translate_source(raw: Any) -> str:
    if raw is None:
        return ""
    s = str(raw).strip()
    if s.startswith("kalman(") and s.endswith(")"):
        inner = s[len("kalman("):-1]
        return "Kalman (" + _SOURCE_BASE.get(inner, inner) + ")"
    base = _SOURCE_BASE.get(s, s)
    return base[:1].upper() + base[1:] if base else ""


def _display_length(value: Any, fmt: Optional[str]) -> int:
    if value is None:
        return 0
    if isinstance(value, datetime):
        return 8
    if isinstance(value, (int, float)) and fmt:
        if fmt in (FMT_M, FMT_S):
            return len(f"{float(value):.3f}")
        if fmt == FMT_DEG:
            return len(f"{float(value):.1f}")
        if fmt == FMT_PCT:
            return len(f"{float(value):.1f}") + 1
        if fmt == FMT_INT:
            return len(f"{int(round(float(value)))}")
    return len(str(value))


def autofit_columns(ws: Worksheet) -> None:
    merged: set[str] = set()
    for cell_range in ws.merged_cells.ranges:
        for row in ws[cell_range.coord]:
            for cell in row:
                merged.add(cell.coordinate)

    widths: dict[str, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None or cell.coordinate in merged:
                continue
            length = _display_length(cell.value, cell.number_format)
            column = cell.column_letter
            if length > widths.get(column, 0):
                widths[column] = length

    for column, length in widths.items():
        ws.column_dimensions[column].width = min(
            max(length + _COL_PADDING, _COL_MIN_WIDTH), _COL_MAX_WIDTH
        )


def section_title(ws: Worksheet, row: int, text: str, span: int = 4) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = LEFT
    ws.row_dimensions[row].height = 20
    return row + 1


def kv(ws: Worksheet, row: int, label: str, value: Any, fmt: Optional[str] = None) -> int:
    label_cell = ws.cell(row=row, column=1, value=label)
    label_cell.font = LABEL_FONT
    label_cell.alignment = LEFT
    value_cell = ws.cell(row=row, column=2, value=value)
    value_cell.font = BASE_FONT
    if fmt is not None:
        value_cell.number_format = fmt
        value_cell.alignment = RIGHT
    else:
        value_cell.alignment = LEFT
    return row + 1


def table_header(ws: Worksheet, row: int, headers: list[str]) -> int:
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_index, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = CELL_BORDER
    return row + 1


def fmt_num_it(value: float, decimals: int = 3) -> str:
    return f"{value:.{decimals}f}".replace(".", ",")
