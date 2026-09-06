from __future__ import annotations

from pathlib import Path
from typing import Optional

from openpyxl import Workbook

from drone.data.flight_excel_sheets import (
    write_autopilot_sheet,
    write_kalman_sheet,
    write_legend_sheet,
    write_parameters_sheet,
    write_summary_sheet,
)
from drone.data.flight_excel_style import fmt_num_it


def collect_run_parameters(logger, app_config, *, path_name: Optional[str] = None) -> list[tuple[str, list[tuple[str, str]]]]:
    ap = app_config.apriltag_autopilot
    pf = app_config.pose_filter
    cp = app_config.camera_pose

    xy_tol = logger.autopilot_xy_tolerance_m if logger.autopilot_xy_tolerance_m is not None else ap.xy_tolerance_m
    z_tol = logger.autopilot_z_tolerance_m if logger.autopilot_z_tolerance_m is not None else ap.z_tolerance_m
    yaw_tol = logger.autopilot_yaw_tolerance_deg if logger.autopilot_yaw_tolerance_deg is not None else ap.yaw_tolerance_deg

    n_waypoints = len({e.get("target_index") for e in logger.autopilot_entries if e.get("target_index") is not None})

    def _n(value, decimals=3):
        return fmt_num_it(float(value), decimals)

    fusione = (
        f"media pesata (esponente {_n(cp.fusion_distance_weight_exponent, 1)})"
        if cp.fusion_mode == "weighted_average" else str(cp.fusion_mode)
    )
    priorita_z = (
        f"ingresso {_n(ap.z_priority_enter_m, 2)} m / uscita {_n(ap.z_priority_exit_m, 2)} m"
        if ap.z_priority_enabled else "disattivata"
    )
    timeout_wp = (
        f"{_n(ap.waypoint_timeout_sec, 0)} s" if ap.waypoint_timeout_enabled else "disattivato"
    )
    gate = (
        f"{_n(pf.outlier_gate_threshold, 1)} (max {pf.outlier_gate_max_consecutive} consecutivi)"
        if pf.outlier_gate_enabled else "disattivato"
    )
    camera = f"calibrata (fx≈{cp.camera_matrix[0][0]:.0f}, fy≈{cp.camera_matrix[1][1]:.0f} px)"

    return [
        ("Missione", [
            ("Percorso", path_name if path_name else "—"),
            ("Waypoint percorsi (distinti)", str(n_waypoints)),
            ("Atterraggio automatico a fine missione", "Sì" if ap.auto_land_on_finish else "No"),
        ]),
        ("Controllo (autopilota proporzionale)", [
            ("Guadagno proporzionale orizzontale (kp XY)", _n(ap.kp_xy, 1)),
            ("Guadagno proporzionale di quota (kp Z)", _n(ap.kp_z, 1)),
            ("Guadagno proporzionale di orientamento (kp yaw)", _n(ap.kp_yaw, 1)),
            ("Saturazione comando orizzontale [canale RC]", str(ap.max_xy_speed)),
            ("Saturazione comando di quota [canale RC]", str(ap.max_z_speed)),
            ("Saturazione comando di rotazione [canale RC]", str(ap.max_yaw_speed)),
            ("Tolleranza orizzontale XY [m]", _n(xy_tol)),
            ("Tolleranza di quota Z [m]", _n(z_tol)),
            ("Tolleranza di orientamento yaw [°]", _n(yaw_tol, 1)),
            ("Priorità di quota (isteresi)", priorita_z),
            ("Timeout posa [s]", _n(ap.pose_timeout_sec, 1)),
            ("Timeout waypoint", timeout_wp),
        ]),
        ("Filtro di Kalman (posizione)", [
            ("Rumore di processo (process noise)", _n(pf.process_noise, 1)),
            ("Rumore di misura (measurement noise)", _n(pf.measurement_noise, 2)),
            ("Gate anti-outlier (soglia)", gate),
            ("Filtro sull'orientamento (yaw)", "attivo" if pf.yaw_filter_enabled else "disattivo"),
        ]),
        ("Localizzazione (AprilTag)", [
            ("Famiglia dei tag", str(cp.tag_family)),
            ("Dimensione del tag [m]", _n(cp.tag_size_m, 2)),
            ("Modalità di fusione dei tag", fusione),
            ("Distanza massima del tag [m]", _n(cp.max_tag_distance_m, 1)),
            ("Camera", camera),
        ]),
        ("Sicurezza batteria", [
            ("Rientro alla base [%]", str(app_config.battery_rth_pct)),
            ("Avviso batteria [%]", str(app_config.battery_warning_pct)),
            ("Atterraggio critico [%]", str(app_config.battery_critical_pct)),
        ]),
    ]

def _build_workbook(logger, parameters: Optional[list[tuple[str, list[tuple[str, str]]]]] = None) -> Workbook:
    separator = logger.TAG_IDS_SEPARATOR
    auto = list(logger.autopilot_entries)
    comp = list(logger.comparison_entries)

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Riepilogo"
    write_summary_sheet(summary_ws, logger)

    if parameters:
        write_parameters_sheet(wb.create_sheet("Parametri"), parameters)

    if auto:
        write_autopilot_sheet(wb.create_sheet("Autopilota"), auto, separator=separator)

    if comp:
        write_kalman_sheet(wb.create_sheet("Kalman"), comp, separator=separator)

    write_legend_sheet(
        wb.create_sheet("Legenda"),
        include_autopilot=bool(auto),
        include_kalman=bool(comp),
    )

    return wb


def save_session_workbook(logger, output_path: Path, parameters=None) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = _build_workbook(logger, parameters)
    wb.save(output_path)
    return output_path
